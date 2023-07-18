import os
import openpyxl
import pathlib
import shutil

def pick_directory(value):
    for genre, videos in genre_dict.items():
        for video in videos:
            if value == video:
                return genre
    return "Others"

if __name__ == '__main__':
    genre_dict = {}
    list_genres = []

    # Pull video IDs and Titles from excel file
    path = 'D:\Tools\Rename-Sort-Files\database.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['main-genres']

    m_row = sheet.max_row

    for i in range(2, m_row):
        genre = sheet.cell(row=i, column=1)
        list_genres.append(genre.value)
        genre_dict[genre.value] = []


    # Match the files with their genre
    path2 = pathlib.Path('D:\\Tools\\Rename-Sort-Files\\test')
    for video in os.listdir(path2):
        for genre, videos in genre_dict.items():
            if genre in video:
                videos.append(video)

    # Create the folders of genres and Move files accordingly
    print("Creating folders of genre and Moving files")
    for f in path2.iterdir():
        if f.is_dir():  # skip folders
            continue
        folder = pick_directory(f.name)
        new_path = os.path.join(path2, folder)
        if os.path.exists(new_path) == False:
            os.mkdir(new_path)
        shutil.move(f, new_path)

    print('Mission completed!!!')
