import openpyxl
import os

if __name__ == '__main__':
    video_titles = []
    video_ids = []
    id_dict = {}

    # Pull video IDs and Titles from excel file
    print('Puling data from excel file ...')
    path = 'D:\Tools\Rename-Sort-Files\database.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['video-test']

    m_row = sheet.max_row

    for i in range(2, m_row + 1):
        title = sheet.cell(row=i, column=1)
        ids = sheet.cell(row=i, column=2)
        video_titles.append(title.value)
        video_ids.append(str(int(ids.value)))
        # Match Ids with Titles
        id_dict[str(int(ids.value))] = title.value
    print('- Finish loading Ids and Titles')

    # Rename files from IDs to Titles
    path2 = 'D:\\Tools\\Rename-Sort-Files\\test'
    os.chdir(path2)

print('Renaming files from IDs to Titles ...')
for f in os.listdir():
    for ids, titles in id_dict.items():
        video_name, extension = os.path.splitext(f)
        if ids == video_name:
            name = f'{titles}{extension}'
            os.rename(f, name)
print('- Finish renaming')